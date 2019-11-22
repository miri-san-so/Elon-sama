from pymongo import MongoClient
from datetime import datetime
from uuid import uuid4
from bson.json_util import loads, dumps
import openpyxl
from openpyxl.styles import Font
import openpyxl
import matplotlib.pyplot as plt

# Gets Current Date
today = str(datetime.now().date())

# Gets Current Time
time = str(datetime.now().time())

# Create Conneciton to the Local MongoDB
client = MongoClient(host="localhost", port=27017)

# List all the databases in the MongoDB
dblist = client.list_database_names()

# If the Database Doesnt Exist Create it
if "elon" not in dblist:
    database = client["elon"]

# Else Select the `elon` database
database = client.elon

# Create if the Collection Does not exist or just select it
collection = database["days"]


def getLastEntry(collection):
    today = str(datetime.now().date())
    query = collection.find({"date": today}, {"date": 1, "_id": 0})
    for i in query:
        return (i['date'])

def insertExcelData(date, task, hour, time):
    #taskList = [date, task, hour]
    # Workbook is created
    try:
        wb = openpyxl.load_workbook("tasks.xlsx")
        sheet = wb.active
    except:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = "date"
        sheet['B1'] = "task"
        sheet['C1'] = "hour"
        sheet['D1'] = "time"        
        ft = Font(name='Arial',size=11,bold=True, italic=False,vertAlign=None,underline='none', strike=False)
        currentCell = sheet['A1']  # or currentCell = ws['A1']
        currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
        currentCell.font = ft
        currentCell = sheet['B1']  # or currentCell = ws['A1']
        currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
        currentCell.font = ft
        currentCell = sheet['C1']  # or currentCell = ws['A1']
        currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
        currentCell.font = ft
        currentCell = sheet['D1']  # or currentCell = ws['A1']
        currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
        currentCell.font = ft
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['A'].width = 10

    row_num = sheet.max_row + 1

    # writing in the excel
    sheet['A'+str(row_num)] = date
    sheet['B'+str(row_num)] = task
    sheet['C'+str(row_num)] = hour
    sheet['D'+str(row_num)] = time
    currentCell = sheet['A'+str(row_num)]
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
    currentCell = sheet['B'+str(row_num)]
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
    currentCell = sheet['C'+str(row_num)]
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
    currentCell = sheet['D'+str(row_num)]
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    wb.save("tasks.xlsx")

def addInSameDay(collection):
    task = input("Enter What task you performed?\n>>> ")
    hours = float(input("How many hours did you take?\n>>> "))

    query = collection.update_one({"date": today}, {'$push': {"day": {"title": task,
                                                                      "hours": hours, "time-created": time}}})
    insertExcelData(today, task, hours, time)
    return(query)

def giveHours(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    # first_sheet = wb.get_sheet_names()[0]
    # worksheet = wb.get_sheet_by_name(first_sheet)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    hours = []

    # here you iterate over the rows in the specific column
    for row in range(2, worksheet.max_row+1):
        if row == lim+2:
            break
        else:
            for column in "C":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                x = worksheet[cell_name].value
                hours.append(x)

    return hours

def giveTime(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    # first_sheet = wb.get_sheet_names()[0]
    # worksheet = wb.get_sheet_by_name(first_sheet)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    dates = []

    # here you iterate over the rows in the specific column
    for row in range(2, worksheet.max_row+1):
        if row == lim+2:
            break
        else:
            for column in "D":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                date = worksheet[cell_name].value
                dates.append(date)

    return dates

def getTitleUsingExcel(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    #first_sheet = first_sheet[0]
    #worksheet = wb.get_sheet_by_name(first_sheet)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]

    hours = []
    
    # here you iterate over the rows in the specific column
    for row in range(2, worksheet.max_row+1):
        if row == lim+2:
            break
        else:
            for column in "B":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                x = worksheet[cell_name].value
                hours.append(x)

    return hours


# insertData() utilizes 
#     > giveLastEntry() to find the last entry
#     > addInSameDay() saves the task in the same day
#     > insertExcelData() saves the data in the excel file called tasks.xlsx
def insertData(collection):
    if getLastEntry(collection) is None:
        newId = uuid4()
        task = input("Enter What task you performed?\n>>> ")
        hours = float(input("How many hours did you take?\n>>> "))
        sample = {"_id": str(newId), "date": today,
                  "day": [{"title": task, "hours": hours, "time-created": time}]}
        x = collection.insert_one(sample)

        print('inserted succesfully ', x)
        insertExcelData(today, task, hours,time)
    else:
        print(addInSameDay(collection))

# giveGraph() utilizes 
#     > giveHours() returns the list of time from the excel sheet
#     > giveTime() returns the list of time from the excel sheet
def giveGraph(lim):
    hoursWorked = giveHours(lim)
    hours = giveTime(lim)
    xAxis = [i + 0.5 for i, _ in enumerate(hours)]

    # Setting the figure size
    plt.figure(figsize=(10, 6))
    ax = plt.axes()
    # Setting the background color
    ax.set_facecolor("#202020")

    plt.bar(xAxis, hoursWorked, color='#AD5CC1')
    plt.title('hoursWorked Vs hours', fontsize=14)
    plt.xlabel('hoursWorked', fontsize=14)
    plt.ylabel('hours', fontsize=14)
    plt.xticks([i + 0.5 for i, _ in enumerate(hours)], hours)
    plt.show()


# print(addInSameDay(collection))
# print(getLastEntry(collection))
insertData(collection)

from pymongo import MongoClient
from datetime import datetime
import openpyxl

today = str(datetime.now().date())


def usingMongo():
    # Create Conneciton to the Local MongoDB
    client = MongoClient(host="localhost", port=27017)

    # List all the databases in the MongoDB
    dblist = client.list_database_names()

    # If the Database Doesnt Exist Create it
    if "elon" not in dblist:
        database = client["elon"]

    # Else Select the `elon` database
    db = client.elon

    # Create if the Collection Does not exist or just select it
    collection = db["days"]

    tasks = collection.find({"date": today}, {"day.title": 1, "_id": 0})


def getTitleUsingExcel(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    #first_sheet = first_sheet[0]
    #worksheet = wb.get_sheet_by_name(first_sheet)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    x = worksheet.max_row

    hours = []
    count = 0
    # here you iterate over the rows in the specific column
    for row in range(x, x-lim, -1):
        if count == lim:
            break
        else:
            for column in "B":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                x = worksheet[cell_name].value
                hours.append(str(x))
        count += 1

    return hours


def getHoursUsingExcel(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    #first_sheet = first_sheet[0]
    #worksheet = wb.get_sheet_by_name(first_sheet)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    x = worksheet.max_row

    hours = []
    count = 0
    # here you iterate over the rows in the specific column
    for row in range(x, x-7, -1):
        if count == 5:
            break
        else:
            for column in "C":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                x = worksheet[cell_name].value
                hours.append(str(x))
        count += 1
    return hours

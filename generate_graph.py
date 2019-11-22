import openpyxl
import matplotlib.pyplot as plt


def giveHours(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)

    hours = []

    # here you iterate over the rows in the specific column
    for row in range(2, worksheet.max_row+1):
        if row == lim+2:
            break
        else:
            for column in "C":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                x = worksheet[cell_name].value
                hours.append(x)

    return hours


def giveDate(lim):
    wb = openpyxl.load_workbook('tasks.xlsx')
    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)

    dates = []

    # here you iterate over the rows in the specific column
    for row in range(2, worksheet.max_row+1):
        if row == lim+2:
            break
        else:
            for column in "D":  # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                # the value of the specific cell
                date = worksheet[cell_name].value
                dates.append(date)

    return dates


def giveGraph(lim):
    hoursWorked = giveHours(lim)
    hours = giveDate(lim)
    xAxis = [i + 0.5 for i, _ in enumerate(hours)]

    # Setting the figure size
    plt.figure(figsize=(10, 6))
    ax = plt.axes()
    # Setting the background color
    ax.set_facecolor("#202020")

    plt.bar(xAxis, hoursWorked, color='#AD5CC1')
    plt.title('Hours Worked Vs hours', fontsize=14)
    plt.xlabel('HoursWorked', fontsize=14)
    plt.ylabel('Hours', fontsize=14)
    plt.xticks([i + 0.5 for i, _ in enumerate(hours)], hours)
    plt.show()

